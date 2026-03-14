import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import BarChart, PieChart, Reference, BarChart3D
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.comments import Comment
from datetime import date, datetime

wb = Workbook()

# ============================================================
# STYLES
# ============================================================
DARK_BLUE = "1B3A5C"
MED_BLUE = "4472C4"
LIGHT_BLUE = "D6E4F0"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
GREEN = "00B050"
YELLOW = "FFC000"
RED = "FF0000"
ORANGE = "ED7D31"
DARK_GREEN = "375623"
LIGHT_GREEN = "E2EFDA"
LIGHT_YELLOW = "FFF2CC"
LIGHT_RED = "FCE4EC"

header_font = Font(name="Arial", bold=True, color=WHITE, size=11)
header_fill = PatternFill("solid", fgColor=DARK_BLUE)
sub_header_font = Font(name="Arial", bold=True, color=WHITE, size=10)
sub_header_fill = PatternFill("solid", fgColor=MED_BLUE)
title_font = Font(name="Arial", bold=True, color=DARK_BLUE, size=16)
subtitle_font = Font(name="Arial", bold=True, color=DARK_BLUE, size=12)
normal_font = Font(name="Arial", size=10)
bold_font = Font(name="Arial", bold=True, size=10)
number_font = Font(name="Arial", size=10)
kpi_value_font = Font(name="Arial", bold=True, color=DARK_BLUE, size=24)
kpi_label_font = Font(name="Arial", size=9, color="666666")
green_font = Font(name="Arial", bold=True, color=GREEN, size=10)
red_font = Font(name="Arial", bold=True, color=RED, size=10)
link_font = Font(name="Arial", color=MED_BLUE, size=10)

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
right_align = Alignment(horizontal='right', vertical='center')
vnd_format = '#,##0'
pct_format = '0.0%'
date_format = 'DD/MM/YYYY'

def style_header_row(ws, row, max_col, fill=None, font=None):
    f = fill or header_fill
    fn = font or header_font
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = fn
        cell.fill = f
        cell.alignment = center_align
        cell.border = thin_border

def style_data_row(ws, row, max_col, alt=False):
    fill = PatternFill("solid", fgColor=LIGHT_GRAY) if alt else PatternFill("solid", fgColor=WHITE)
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = normal_font
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = left_align

def auto_width(ws, max_col, min_width=12, max_width=35):
    for col in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = min_width

# ============================================================
# SHEET: Config
# ============================================================
ws_config = wb.active
ws_config.title = "Config"
ws_config.sheet_properties.tabColor = "808080"

ws_config['A1'] = "CẤU HÌNH HỆ THỐNG"
ws_config['A1'].font = title_font
ws_config.merge_cells('A1:D1')

# Months
ws_config['A3'] = "Tháng"
ws_config['A3'].font = bold_font
months = ["Tháng 1","Tháng 2","Tháng 3","Tháng 4","Tháng 5","Tháng 6",
          "Tháng 7","Tháng 8","Tháng 9","Tháng 10","Tháng 11","Tháng 12"]
for i, m in enumerate(months):
    ws_config.cell(row=4+i, column=1, value=m)

# Years
ws_config['B3'] = "Năm"
ws_config['B3'].font = bold_font
for i, y in enumerate(range(2024, 2031)):
    ws_config.cell(row=4+i, column=2, value=y)

# Status
ws_config['C3'] = "Trạng Thái Dự Án"
ws_config['C3'].font = bold_font
statuses = ["On Track", "At Risk", "Delayed", "Completed", "On Hold"]
for i, s in enumerate(statuses):
    ws_config.cell(row=4+i, column=3, value=s)

# Priority
ws_config['D3'] = "Mức Ưu Tiên"
ws_config['D3'].font = bold_font
priorities = ["Critical", "High", "Medium", "Low"]
for i, p in enumerate(priorities):
    ws_config.cell(row=4+i, column=4, value=p)

# Risk severity
ws_config['E3'] = "Mức Nghiêm Trọng"
ws_config['E3'].font = bold_font
severities = ["Critical", "High", "Medium", "Low"]
for i, s in enumerate(severities):
    ws_config.cell(row=4+i, column=5, value=s)

# Risk status
ws_config['F3'] = "Trạng Thái Rủi Ro"
ws_config['F3'].font = bold_font
risk_statuses = ["Open", "In Progress", "Resolved", "Closed", "Escalated"]
for i, s in enumerate(risk_statuses):
    ws_config.cell(row=4+i, column=6, value=s)

# Probability
ws_config['G3'] = "Xác Suất"
ws_config['G3'].font = bold_font
probs = ["Rất cao (90%)", "Cao (70%)", "Trung bình (50%)", "Thấp (30%)", "Rất thấp (10%)"]
for i, p in enumerate(probs):
    ws_config.cell(row=4+i, column=7, value=p)

# Payment status
ws_config['H3'] = "TT Thanh Toán"
ws_config['H3'].font = bold_font
pay_statuses = ["Đã thanh toán", "Chờ thanh toán", "Quá hạn", "Chưa đến hạn"]
for i, s in enumerate(pay_statuses):
    ws_config.cell(row=4+i, column=8, value=s)

# Scoring tables
ws_config['A18'] = "BẢNG QUY ĐỔI ĐIỂM RỦI RO"
ws_config['A18'].font = subtitle_font
ws_config['A19'] = "Mức"; ws_config['B19'] = "Điểm"
ws_config['A19'].font = bold_font; ws_config['B19'].font = bold_font
score_map = [("Critical", 4), ("High", 3), ("Medium", 2), ("Low", 1)]
for i, (label, score) in enumerate(score_map):
    ws_config.cell(row=20+i, column=1, value=label)
    ws_config.cell(row=20+i, column=2, value=score)

style_header_row(ws_config, 3, 8)
for col in range(1, 9):
    ws_config.column_dimensions[get_column_letter(col)].width = 18

# ============================================================
# SHEET: Danh Sách Dự Án
# ============================================================
ws_projects = wb.create_sheet("Danh Sách Dự Án")
ws_projects.sheet_properties.tabColor = MED_BLUE

ws_projects['A1'] = "BÁO CÁO TỔNG HỢP DỰ ÁN - TIÊN PHONG CDS"
ws_projects['A1'].font = title_font
ws_projects.merge_cells('A1:P1')
ws_projects['A2'] = '=IF(Dashboard!B3<>"",Dashboard!B3&" - "&Dashboard!B4,"")'
ws_projects['A2'].font = subtitle_font
ws_projects.merge_cells('A2:P2')

headers = ["STT", "Mã DA", "Tên Dự Án", "Khách Hàng", "PM Phụ Trách",
           "Ngày BĐ", "Ngày KT (KH)", "Ngày KT (TT)", "Trạng Thái",
           "% HT Kế Hoạch", "% HT Thực Tế", "Chênh Lệch",
           "Giá Trị HĐ (VNĐ)", "Đã Thanh Toán (VNĐ)", "Còn Lại (VNĐ)", "% Thanh Toán",
           "Số Nhân Sự", "Mức Ưu Tiên", "Ghi Chú"]
for i, h in enumerate(headers, 1):
    ws_projects.cell(row=4, column=i, value=h)
style_header_row(ws_projects, 4, len(headers))

# Sample data
sample_projects = [
    [1, "CDS-2026-001", "CDS Platform - Module Sales", "VNFM Holdings", "Nguyễn Văn Minh",
     date(2025,11,1), date(2026,4,30), None, "On Track",
     None, 0.65, None, 1200000000, 480000000, None, None, 5, "High", "Phase 2 đang triển khai"],
    [2, "CDS-2026-002", "ERP Integration Gateway", "Nova Group", "Trần Thị Hương",
     date(2025,12,15), date(2026,6,30), None, "At Risk",
     None, 0.40, None, 850000000, 255000000, None, None, 4, "Critical", "Delay do API vendor"],
    [3, "CDS-2026-003", "Data Analytics Dashboard", "Phú Mỹ Hưng", "Lê Quốc Bảo",
     date(2026,1,10), date(2026,5,15), None, "On Track",
     None, 0.55, None, 650000000, 325000000, None, None, 3, "Medium", ""],
    [4, "CDS-2026-004", "HR Management System", "Tập đoàn Thành Công", "Phạm Đức Anh",
     date(2025,10,1), date(2026,3,31), date(2026,3,20), "Completed",
     None, 1.0, None, 500000000, 500000000, None, None, 3, "Medium", "Đã nghiệm thu"],
    [5, "CDS-2026-005", "Supply Chain Optimization", "Vinamilk", "Nguyễn Thị Lan",
     date(2026,2,1), date(2026,8,31), None, "Delayed",
     None, 0.15, None, 2100000000, 420000000, None, None, 7, "Critical", "Thiếu BA, cần tăng cường"],
    [6, "CDS-2026-006", "Mobile App - Customer Portal", "FPT Retail", "Hoàng Minh Tuấn",
     date(2026,1,15), date(2026,7,15), None, "On Track",
     None, 0.35, None, 780000000, 234000000, None, None, 4, "High", "Sprint 3/8"],
    [7, "CDS-2026-007", "AI Chatbot Integration", "Techcombank", "Võ Thanh Tùng",
     date(2026,3,1), date(2026,9,30), None, "On Track",
     None, 0.05, None, 1500000000, 300000000, None, None, 6, "High", "Vừa kick-off"],
]

for idx, proj in enumerate(sample_projects):
    row = 5 + idx
    for col, val in enumerate(proj, 1):
        cell = ws_projects.cell(row=row, column=col, value=val)
    style_data_row(ws_projects, row, len(headers), alt=(idx % 2 == 1))

    # Date format
    for dc in [6, 7, 8]:
        ws_projects.cell(row=row, column=dc).number_format = date_format

    # Formula: % HT Kế Hoạch = (TODAY - Start) / (End - Start)
    ws_projects.cell(row=row, column=10).value = f'=IF(I{row}="Completed",1,MIN(1,MAX(0,(TODAY()-F{row})/(G{row}-F{row}))))'
    ws_projects.cell(row=row, column=10).number_format = pct_format

    # % HT Thực Tế already set
    ws_projects.cell(row=row, column=11).number_format = pct_format

    # Chênh lệch = Thực tế - Kế hoạch
    ws_projects.cell(row=row, column=12).value = f'=K{row}-J{row}'
    ws_projects.cell(row=row, column=12).number_format = '+0.0%;-0.0%;0.0%'

    # VND format
    ws_projects.cell(row=row, column=13).number_format = vnd_format
    ws_projects.cell(row=row, column=14).number_format = vnd_format

    # Còn lại = HĐ - Đã TT
    ws_projects.cell(row=row, column=15).value = f'=M{row}-N{row}'
    ws_projects.cell(row=row, column=15).number_format = vnd_format

    # % Thanh toán
    ws_projects.cell(row=row, column=16).value = f'=IF(M{row}=0,0,N{row}/M{row})'
    ws_projects.cell(row=row, column=16).number_format = pct_format

# Totals row
total_row = 5 + len(sample_projects)
ws_projects.cell(row=total_row, column=1, value="TỔNG CỘNG")
ws_projects.cell(row=total_row, column=1).font = bold_font
ws_projects.merge_cells(f'A{total_row}:D{total_row}')
for tc in [13, 14, 15]:
    col_letter = get_column_letter(tc)
    ws_projects.cell(row=total_row, column=tc).value = f'=SUM({col_letter}5:{col_letter}{total_row-1})'
    ws_projects.cell(row=total_row, column=tc).number_format = vnd_format
    ws_projects.cell(row=total_row, column=tc).font = bold_font

ws_projects.cell(row=total_row, column=10).value = f'=AVERAGE(J5:J{total_row-1})'
ws_projects.cell(row=total_row, column=10).number_format = pct_format
ws_projects.cell(row=total_row, column=10).font = bold_font
ws_projects.cell(row=total_row, column=11).value = f'=AVERAGE(K5:K{total_row-1})'
ws_projects.cell(row=total_row, column=11).number_format = pct_format
ws_projects.cell(row=total_row, column=11).font = bold_font
ws_projects.cell(row=total_row, column=16).value = f'=IF(M{total_row}=0,0,N{total_row}/M{total_row})'
ws_projects.cell(row=total_row, column=16).number_format = pct_format
ws_projects.cell(row=total_row, column=16).font = bold_font
ws_projects.cell(row=total_row, column=17).value = f'=SUM(Q5:Q{total_row-1})'
ws_projects.cell(row=total_row, column=17).font = bold_font

style_header_row(ws_projects, total_row, len(headers), fill=PatternFill("solid", fgColor=LIGHT_BLUE), font=Font(name="Arial", bold=True, color=DARK_BLUE, size=10))

# Data validations
dv_status = DataValidation(type="list", formula1="Config!$C$4:$C$8", allow_blank=True)
dv_status.error = "Chọn trạng thái từ danh sách"
dv_status.errorTitle = "Giá trị không hợp lệ"
ws_projects.add_data_validation(dv_status)
dv_status.add(f'I5:I{total_row-1}')

dv_priority = DataValidation(type="list", formula1="Config!$D$4:$D$7", allow_blank=True)
ws_projects.add_data_validation(dv_priority)
dv_priority.add(f'R5:R{total_row-1}')

# Conditional formatting for status
ws_projects.conditional_formatting.add(f'I5:I{total_row-1}',
    CellIsRule(operator='equal', formula=['"On Track"'], fill=PatternFill("solid", fgColor=LIGHT_GREEN)))
ws_projects.conditional_formatting.add(f'I5:I{total_row-1}',
    CellIsRule(operator='equal', formula=['"At Risk"'], fill=PatternFill("solid", fgColor=LIGHT_YELLOW)))
ws_projects.conditional_formatting.add(f'I5:I{total_row-1}',
    CellIsRule(operator='equal', formula=['"Delayed"'], fill=PatternFill("solid", fgColor=LIGHT_RED)))
ws_projects.conditional_formatting.add(f'I5:I{total_row-1}',
    CellIsRule(operator='equal', formula=['"Completed"'], fill=PatternFill("solid", fgColor="C6EFCE")))

# Conditional formatting for chênh lệch
ws_projects.conditional_formatting.add(f'L5:L{total_row-1}',
    CellIsRule(operator='lessThan', formula=['0'], font=Font(color=RED, bold=True), fill=PatternFill("solid", fgColor=LIGHT_RED)))
ws_projects.conditional_formatting.add(f'L5:L{total_row-1}',
    CellIsRule(operator='greaterThanOrEqual', formula=['0'], font=Font(color=DARK_GREEN, bold=True), fill=PatternFill("solid", fgColor=LIGHT_GREEN)))

# Column widths
col_widths = [5, 14, 28, 20, 18, 13, 13, 13, 12, 13, 13, 12, 18, 18, 18, 13, 10, 12, 25]
for i, w in enumerate(col_widths, 1):
    ws_projects.column_dimensions[get_column_letter(i)].width = w

ws_projects.freeze_panes = 'A5'
ws_projects.sheet_view.zoomScale = 85

# ============================================================
# SHEET: Tiến Độ Hợp Đồng
# ============================================================
ws_contract = wb.create_sheet("Tiến Độ Hợp Đồng")
ws_contract.sheet_properties.tabColor = GREEN

ws_contract['A1'] = "THEO DÕI TIẾN ĐỘ THANH TOÁN HỢP ĐỒNG"
ws_contract['A1'].font = title_font
ws_contract.merge_cells('A1:T1')

headers_c = ["STT", "Mã DA", "Tên Dự Án", "Khách Hàng", "Giá Trị HĐ (VNĐ)",
             "Đợt 1 - Số Tiền", "Đợt 1 - Ngày DK", "Đợt 1 - Ngày TT", "Đợt 1 - TT",
             "Đợt 2 - Số Tiền", "Đợt 2 - Ngày DK", "Đợt 2 - Ngày TT", "Đợt 2 - TT",
             "Đợt 3 - Số Tiền", "Đợt 3 - Ngày DK", "Đợt 3 - Ngày TT", "Đợt 3 - TT",
             "Tổng Đã Thu (VNĐ)", "Tổng Còn Thu (VNĐ)", "% Thu"]

for i, h in enumerate(headers_c, 1):
    ws_contract.cell(row=3, column=i, value=h)
style_header_row(ws_contract, 3, len(headers_c))

# Sub-headers for payment batches
for batch_start in [6, 10, 14]:
    for col in range(batch_start, batch_start + 4):
        ws_contract.cell(row=3, column=col).fill = sub_header_fill

# Sample payment data
payment_data = [
    [1, "CDS-2026-001", "CDS Platform - Module Sales", "VNFM Holdings", 1200000000,
     360000000, date(2025,11,15), date(2025,11,20), "Đã thanh toán",
     120000000, date(2026,2,28), date(2026,3,1), "Đã thanh toán",
     360000000, date(2026,4,30), None, "Chưa đến hạn"],
    [2, "CDS-2026-002", "ERP Integration Gateway", "Nova Group", 850000000,
     255000000, date(2025,12,20), date(2025,12,25), "Đã thanh toán",
     255000000, date(2026,3,31), None, "Chờ thanh toán",
     340000000, date(2026,6,30), None, "Chưa đến hạn"],
    [3, "CDS-2026-003", "Data Analytics Dashboard", "Phú Mỹ Hưng", 650000000,
     195000000, date(2026,1,15), date(2026,1,18), "Đã thanh toán",
     130000000, date(2026,3,15), date(2026,3,10), "Đã thanh toán",
     325000000, date(2026,5,15), None, "Chưa đến hạn"],
    [4, "CDS-2026-004", "HR Management System", "Tập đoàn Thành Công", 500000000,
     150000000, date(2025,10,10), date(2025,10,12), "Đã thanh toán",
     150000000, date(2026,1,15), date(2026,1,15), "Đã thanh toán",
     200000000, date(2026,3,31), date(2026,3,20), "Đã thanh toán"],
    [5, "CDS-2026-005", "Supply Chain Optimization", "Vinamilk", 2100000000,
     420000000, date(2026,2,10), date(2026,2,15), "Đã thanh toán",
     630000000, date(2026,5,31), None, "Chưa đến hạn",
     1050000000, date(2026,8,31), None, "Chưa đến hạn"],
    [6, "CDS-2026-006", "Mobile App - Customer Portal", "FPT Retail", 780000000,
     234000000, date(2026,1,20), date(2026,1,22), "Đã thanh toán",
     234000000, date(2026,4,15), None, "Chờ thanh toán",
     312000000, date(2026,7,15), None, "Chưa đến hạn"],
    [7, "CDS-2026-007", "AI Chatbot Integration", "Techcombank", 1500000000,
     300000000, date(2026,3,5), date(2026,3,8), "Đã thanh toán",
     450000000, date(2026,6,30), None, "Chưa đến hạn",
     750000000, date(2026,9,30), None, "Chưa đến hạn"],
]

for idx, data in enumerate(payment_data):
    row = 4 + idx
    for col, val in enumerate(data, 1):
        ws_contract.cell(row=row, column=col, value=val)
    style_data_row(ws_contract, row, len(headers_c), alt=(idx % 2 == 1))

    ws_contract.cell(row=row, column=5).number_format = vnd_format
    for vc in [6, 10, 14]:
        ws_contract.cell(row=row, column=vc).number_format = vnd_format
    for dc in [7, 8, 11, 12, 15, 16]:
        ws_contract.cell(row=row, column=dc).number_format = date_format

    # Tổng đã thu = SUM of paid amounts
    ws_contract.cell(row=row, column=18).value = (
        f'=SUMPRODUCT((I{row}="Đã thanh toán")*F{row})+SUMPRODUCT((M{row}="Đã thanh toán")*J{row})+SUMPRODUCT((Q{row}="Đã thanh toán")*N{row})'
    )
    ws_contract.cell(row=row, column=18).number_format = vnd_format

    # Tổng còn thu
    ws_contract.cell(row=row, column=19).value = f'=E{row}-R{row}'
    ws_contract.cell(row=row, column=19).number_format = vnd_format

    # % Thu
    ws_contract.cell(row=row, column=20).value = f'=IF(E{row}=0,0,R{row}/E{row})'
    ws_contract.cell(row=row, column=20).number_format = pct_format

# Totals
tr = 4 + len(payment_data)
ws_contract.cell(row=tr, column=1, value="TỔNG CỘNG").font = bold_font
ws_contract.merge_cells(f'A{tr}:D{tr}')
for tc in [5, 6, 10, 14, 18, 19]:
    cl = get_column_letter(tc)
    ws_contract.cell(row=tr, column=tc).value = f'=SUM({cl}4:{cl}{tr-1})'
    ws_contract.cell(row=tr, column=tc).number_format = vnd_format
    ws_contract.cell(row=tr, column=tc).font = bold_font
ws_contract.cell(row=tr, column=20).value = f'=IF(E{tr}=0,0,R{tr}/E{tr})'
ws_contract.cell(row=tr, column=20).number_format = pct_format
ws_contract.cell(row=tr, column=20).font = bold_font

style_header_row(ws_contract, tr, len(headers_c), fill=PatternFill("solid", fgColor=LIGHT_BLUE), font=Font(name="Arial", bold=True, color=DARK_BLUE, size=10))

# Data validation payment status
dv_pay = DataValidation(type="list", formula1="Config!$H$4:$H$7", allow_blank=True)
ws_contract.add_data_validation(dv_pay)
for pc in [9, 13, 17]:
    dv_pay.add(f'{get_column_letter(pc)}4:{get_column_letter(pc)}{tr-1}')

# Conditional formatting for payment status
for pc in [9, 13, 17]:
    col_range = f'{get_column_letter(pc)}4:{get_column_letter(pc)}{tr-1}'
    ws_contract.conditional_formatting.add(col_range,
        CellIsRule(operator='equal', formula=['"Đã thanh toán"'], fill=PatternFill("solid", fgColor=LIGHT_GREEN)))
    ws_contract.conditional_formatting.add(col_range,
        CellIsRule(operator='equal', formula=['"Quá hạn"'], fill=PatternFill("solid", fgColor=LIGHT_RED)))
    ws_contract.conditional_formatting.add(col_range,
        CellIsRule(operator='equal', formula=['"Chờ thanh toán"'], fill=PatternFill("solid", fgColor=LIGHT_YELLOW)))

# Data bar for % Thu
ws_contract.conditional_formatting.add(f'T4:T{tr-1}',
    DataBarRule(start_type='num', start_value=0, end_type='num', end_value=1, color=MED_BLUE))

# Chart: Stacked bar for payment progress
chart_pay = BarChart()
chart_pay.type = "bar"
chart_pay.grouping = "stacked"
chart_pay.title = "Tiến Độ Thanh Toán Theo Dự Án"
chart_pay.y_axis.title = "VNĐ"
chart_pay.style = 10
chart_pay.width = 25
chart_pay.height = 14

cats = Reference(ws_contract, min_col=3, min_row=4, max_row=tr-1)
data_paid = Reference(ws_contract, min_col=18, min_row=3, max_row=tr-1)
data_remain = Reference(ws_contract, min_col=19, min_row=3, max_row=tr-1)
chart_pay.add_data(data_paid, titles_from_data=True)
chart_pay.add_data(data_remain, titles_from_data=True)
chart_pay.set_categories(cats)
chart_pay.series[0].graphicalProperties.solidFill = GREEN
chart_pay.series[1].graphicalProperties.solidFill = ORANGE
ws_contract.add_chart(chart_pay, f"A{tr+2}")

col_widths_c = [5, 14, 28, 20, 18, 15, 13, 13, 14, 15, 13, 13, 14, 15, 13, 13, 14, 18, 18, 10]
for i, w in enumerate(col_widths_c, 1):
    ws_contract.column_dimensions[get_column_letter(i)].width = w

ws_contract.freeze_panes = 'A4'

# ============================================================
# SHEET: Nhân Sự Dự Án
# ============================================================
ws_hr = wb.create_sheet("Nhân Sự Dự Án")
ws_hr.sheet_properties.tabColor = ORANGE

ws_hr['A1'] = "PHÂN BỔ NHÂN SỰ DỰ ÁN"
ws_hr['A1'].font = title_font
ws_hr.merge_cells('A1:J1')

headers_hr = ["STT", "Họ Tên", "Vị Trí", "Dự Án", "Vai Trò Trong DA",
              "Giờ Phân Bổ/Tuần", "Giờ Thực Tế/Tuần", "Utilization Rate",
              "Trạng Thái", "Ghi Chú"]
for i, h in enumerate(headers_hr, 1):
    ws_hr.cell(row=3, column=i, value=h)
style_header_row(ws_hr, 3, len(headers_hr))

staff_data = [
    [1, "Nguyễn Văn Minh", "Senior PM", "CDS-2026-001", "Project Manager", 40, 42, None, "Active", "Lead PM"],
    [2, "Trần Thị Hương", "PM", "CDS-2026-002", "Project Manager", 40, 45, None, "Overloaded", "Cần hỗ trợ"],
    [3, "Lê Quốc Bảo", "Tech Lead", "CDS-2026-003", "Tech Lead", 32, 35, None, "Active", ""],
    [4, "Phạm Đức Anh", "Senior Dev", "CDS-2026-004", "Tech Lead", 40, 38, None, "Available", "DA đã hoàn thành"],
    [5, "Nguyễn Thị Lan", "PM", "CDS-2026-005", "Project Manager", 40, 48, None, "Overloaded", "Thiếu resource"],
    [6, "Hoàng Minh Tuấn", "Mobile Lead", "CDS-2026-006", "Tech Lead", 40, 40, None, "Active", ""],
    [7, "Võ Thanh Tùng", "AI Engineer", "CDS-2026-007", "Tech Lead", 40, 36, None, "Active", "Mới join"],
    [8, "Đỗ Hải Nam", "Backend Dev", "CDS-2026-001", "Developer", 32, 30, None, "Active", "Part-time DA khác"],
    [9, "Trịnh Minh Châu", "Frontend Dev", "CDS-2026-006", "Developer", 40, 38, None, "Active", ""],
    [10, "Ngô Quang Huy", "BA", "CDS-2026-002", "Business Analyst", 40, 44, None, "Overloaded", "Hỗ trợ thêm DA005"],
    [11, "Lý Thị Thu", "QA Lead", "CDS-2026-001", "QA Engineer", 24, 28, None, "Active", "QA cho 2 DA"],
    [12, "Bùi Văn Khoa", "DevOps", "CDS-2026-005", "DevOps Engineer", 32, 35, None, "Active", "Infra setup"],
]

for idx, data in enumerate(staff_data):
    row = 4 + idx
    for col, val in enumerate(data, 1):
        ws_hr.cell(row=row, column=col, value=val)
    style_data_row(ws_hr, row, len(headers_hr), alt=(idx % 2 == 1))

    # Utilization Rate formula
    ws_hr.cell(row=row, column=8).value = f'=IF(F{row}=0,0,G{row}/F{row})'
    ws_hr.cell(row=row, column=8).number_format = pct_format

hr_total = 4 + len(staff_data)
ws_hr.cell(row=hr_total, column=1, value="TRUNG BÌNH").font = bold_font
ws_hr.merge_cells(f'A{hr_total}:E{hr_total}')
ws_hr.cell(row=hr_total, column=6).value = f'=AVERAGE(F4:F{hr_total-1})'
ws_hr.cell(row=hr_total, column=6).font = bold_font
ws_hr.cell(row=hr_total, column=7).value = f'=AVERAGE(G4:G{hr_total-1})'
ws_hr.cell(row=hr_total, column=7).font = bold_font
ws_hr.cell(row=hr_total, column=8).value = f'=AVERAGE(H4:H{hr_total-1})'
ws_hr.cell(row=hr_total, column=8).number_format = pct_format
ws_hr.cell(row=hr_total, column=8).font = bold_font
style_header_row(ws_hr, hr_total, len(headers_hr), fill=PatternFill("solid", fgColor=LIGHT_BLUE), font=Font(name="Arial", bold=True, color=DARK_BLUE, size=10))

# Conditional formatting utilization
ws_hr.conditional_formatting.add(f'H4:H{hr_total-1}',
    CellIsRule(operator='greaterThan', formula=['1.1'], fill=PatternFill("solid", fgColor=LIGHT_RED), font=Font(color=RED, bold=True)))
ws_hr.conditional_formatting.add(f'H4:H{hr_total-1}',
    CellIsRule(operator='between', formula=['0.8', '1.1'], fill=PatternFill("solid", fgColor=LIGHT_GREEN), font=Font(color=DARK_GREEN, bold=True)))
ws_hr.conditional_formatting.add(f'H4:H{hr_total-1}',
    CellIsRule(operator='lessThan', formula=['0.8'], fill=PatternFill("solid", fgColor=LIGHT_YELLOW)))

# Data bar for utilization
ws_hr.conditional_formatting.add(f'H4:H{hr_total-1}',
    DataBarRule(start_type='num', start_value=0, end_type='num', end_value=1.5, color=MED_BLUE))

# Chart: Utilization Rate
chart_util = BarChart()
chart_util.type = "col"
chart_util.title = "Utilization Rate Theo Nhân Sự"
chart_util.y_axis.title = "Utilization Rate"
chart_util.y_axis.scaling.max = 1.5
chart_util.style = 10
chart_util.width = 25
chart_util.height = 14

cats_hr = Reference(ws_hr, min_col=2, min_row=4, max_row=hr_total-1)
data_util = Reference(ws_hr, min_col=8, min_row=3, max_row=hr_total-1)
chart_util.add_data(data_util, titles_from_data=True)
chart_util.set_categories(cats_hr)
chart_util.series[0].graphicalProperties.solidFill = MED_BLUE
ws_hr.add_chart(chart_util, f"A{hr_total+2}")

# MATRIX section
matrix_start = hr_total + 20
ws_hr.cell(row=matrix_start, column=1, value="MA TRẬN NHÂN SỰ x DỰ ÁN").font = subtitle_font
ws_hr.merge_cells(f'A{matrix_start}:H{matrix_start}')

project_codes = ["CDS-2026-001", "CDS-2026-002", "CDS-2026-003", "CDS-2026-004", "CDS-2026-005", "CDS-2026-006", "CDS-2026-007"]
staff_names = list(set([d[1] for d in staff_data]))
staff_names.sort()

ws_hr.cell(row=matrix_start+1, column=1, value="Nhân Sự \\ Dự Án").font = bold_font
for j, pc in enumerate(project_codes):
    ws_hr.cell(row=matrix_start+1, column=2+j, value=pc).font = sub_header_font
    ws_hr.cell(row=matrix_start+1, column=2+j).fill = sub_header_fill
    ws_hr.cell(row=matrix_start+1, column=2+j).alignment = center_align
ws_hr.cell(row=matrix_start+1, column=1).fill = header_fill
ws_hr.cell(row=matrix_start+1, column=1).font = header_font

for i, name in enumerate(staff_names):
    row = matrix_start + 2 + i
    ws_hr.cell(row=row, column=1, value=name).font = bold_font
    for j, pc in enumerate(project_codes):
        matched = [d for d in staff_data if d[1] == name and d[3] == pc]
        if matched:
            ws_hr.cell(row=row, column=2+j, value=matched[0][6])  # hours
            ws_hr.cell(row=row, column=2+j).fill = PatternFill("solid", fgColor=LIGHT_GREEN)
        else:
            ws_hr.cell(row=row, column=2+j, value="-")
        ws_hr.cell(row=row, column=2+j).alignment = center_align
        ws_hr.cell(row=row, column=2+j).border = thin_border
    style_data_row(ws_hr, row, 1+len(project_codes), alt=(i % 2 == 1))

col_widths_hr = [5, 22, 16, 16, 18, 16, 16, 15, 14, 25]
for i, w in enumerate(col_widths_hr, 1):
    ws_hr.column_dimensions[get_column_letter(i)].width = w
ws_hr.freeze_panes = 'A4'

# ============================================================
# SHEET: Rủi Ro & Vấn Đề
# ============================================================
ws_risk = wb.create_sheet("Rủi Ro & Vấn Đề")
ws_risk.sheet_properties.tabColor = RED

ws_risk['A1'] = "QUẢN LÝ RỦI RO & VẤN ĐỀ DỰ ÁN"
ws_risk['A1'].font = title_font
ws_risk.merge_cells('A1:N1')

headers_r = ["STT", "Mã DA", "Tên Dự Án", "Mô Tả Rủi Ro / Vấn Đề",
             "Mức Nghiêm Trọng", "Xác Suất", "Mức Ảnh Hưởng",
             "Điểm Severity", "Điểm Probability", "Điểm Impact",
             "Risk Score", "Người Chịu TN", "Hành Động Giảm Thiểu",
             "Deadline", "Trạng Thái"]
for i, h in enumerate(headers_r, 1):
    ws_risk.cell(row=3, column=i, value=h)
style_header_row(ws_risk, 3, len(headers_r))

risk_data = [
    [1, "CDS-2026-002", "ERP Integration Gateway", "API vendor thay đổi spec không thông báo", "Critical", "Cao (70%)", "High", None, None, None, None, "Trần Thị Hương", "Escalate vendor, backup plan API wrapper", date(2026,3,20), "Open"],
    [2, "CDS-2026-005", "Supply Chain Optimization", "Thiếu BA có kinh nghiệm supply chain", "High", "Rất cao (90%)", "High", None, None, None, None, "Nguyễn Thị Lan", "Tuyển BA freelance, training nội bộ", date(2026,3,25), "In Progress"],
    [3, "CDS-2026-005", "Supply Chain Optimization", "Budget có thể vượt 15% do scope creep", "High", "Cao (70%)", "Critical", None, None, None, None, "Nguyễn Thị Lan", "Change request formal, freeze scope Phase 1", date(2026,4,1), "Open"],
    [4, "CDS-2026-001", "CDS Platform - Module Sales", "Performance bottleneck khi load >1000 concurrent users", "Medium", "Trung bình (50%)", "High", None, None, None, None, "Nguyễn Văn Minh", "Load testing, optimize query, caching", date(2026,4,15), "In Progress"],
    [5, "CDS-2026-006", "Mobile App - Customer Portal", "iOS review policy thay đổi có thể delay release", "Medium", "Thấp (30%)", "Medium", None, None, None, None, "Hoàng Minh Tuấn", "Submit sớm, prepare documentation đầy đủ", date(2026,5,1), "Open"],
    [6, "CDS-2026-007", "AI Chatbot Integration", "Model AI response chưa đạt accuracy 95% target", "High", "Trung bình (50%)", "High", None, None, None, None, "Võ Thanh Tùng", "Fine-tune model, expand training data", date(2026,6,15), "Open"],
    [7, "CDS-2026-002", "ERP Integration Gateway", "Key developer có thể nghỉ việc", "Critical", "Trung bình (50%)", "Critical", None, None, None, None, "Trần Thị Hương", "Knowledge transfer, documentation, retention plan", date(2026,3,31), "Escalated"],
    [8, "CDS-2026-003", "Data Analytics Dashboard", "Data quality từ source system không ổn định", "Medium", "Cao (70%)", "Medium", None, None, None, None, "Lê Quốc Bảo", "Data validation layer, ETL error handling", date(2026,4,10), "In Progress"],
]

for idx, data in enumerate(risk_data):
    row = 4 + idx
    for col, val in enumerate(data, 1):
        ws_risk.cell(row=row, column=col, value=val)
    style_data_row(ws_risk, row, len(headers_r), alt=(idx % 2 == 1))

    ws_risk.cell(row=row, column=14).number_format = date_format

    # Severity score lookup
    ws_risk.cell(row=row, column=8).value = f'=IFERROR(VLOOKUP(E{row},Config!$A$20:$B$23,2,FALSE),0)'
    # Probability score - extract from text
    ws_risk.cell(row=row, column=9).value = f'=IF(ISNUMBER(SEARCH("Rất cao",F{row})),5,IF(ISNUMBER(SEARCH("Cao",F{row})),4,IF(ISNUMBER(SEARCH("Trung bình",F{row})),3,IF(ISNUMBER(SEARCH("Thấp",F{row})),2,IF(ISNUMBER(SEARCH("Rất thấp",F{row})),1,0)))))'
    # Impact score lookup
    ws_risk.cell(row=row, column=10).value = f'=IFERROR(VLOOKUP(G{row},Config!$A$20:$B$23,2,FALSE),0)'
    # Risk Score = Severity × Probability × Impact
    ws_risk.cell(row=row, column=11).value = f'=H{row}*I{row}*J{row}'
    ws_risk.cell(row=row, column=11).font = bold_font

# Risk score conditional formatting
risk_end = 4 + len(risk_data) - 1
ws_risk.conditional_formatting.add(f'K4:K{risk_end}',
    CellIsRule(operator='greaterThanOrEqual', formula=['48'], fill=PatternFill("solid", fgColor="FF0000"), font=Font(color=WHITE, bold=True)))
ws_risk.conditional_formatting.add(f'K4:K{risk_end}',
    CellIsRule(operator='between', formula=['24', '47'], fill=PatternFill("solid", fgColor=ORANGE), font=Font(bold=True)))
ws_risk.conditional_formatting.add(f'K4:K{risk_end}',
    CellIsRule(operator='between', formula=['8', '23'], fill=PatternFill("solid", fgColor=YELLOW), font=Font(bold=True)))
ws_risk.conditional_formatting.add(f'K4:K{risk_end}',
    CellIsRule(operator='lessThan', formula=['8'], fill=PatternFill("solid", fgColor=LIGHT_GREEN)))

# Severity conditional formatting
ws_risk.conditional_formatting.add(f'E4:E{risk_end}',
    CellIsRule(operator='equal', formula=['"Critical"'], fill=PatternFill("solid", fgColor=LIGHT_RED), font=Font(color=RED, bold=True)))
ws_risk.conditional_formatting.add(f'E4:E{risk_end}',
    CellIsRule(operator='equal', formula=['"High"'], fill=PatternFill("solid", fgColor=LIGHT_YELLOW), font=Font(color=ORANGE, bold=True)))

# Status conditional formatting
ws_risk.conditional_formatting.add(f'O4:O{risk_end}',
    CellIsRule(operator='equal', formula=['"Open"'], fill=PatternFill("solid", fgColor=LIGHT_RED)))
ws_risk.conditional_formatting.add(f'O4:O{risk_end}',
    CellIsRule(operator='equal', formula=['"Escalated"'], fill=PatternFill("solid", fgColor="FF0000"), font=Font(color=WHITE, bold=True)))
ws_risk.conditional_formatting.add(f'O4:O{risk_end}',
    CellIsRule(operator='equal', formula=['"Resolved"'], fill=PatternFill("solid", fgColor=LIGHT_GREEN)))

# Data validations
dv_sev = DataValidation(type="list", formula1="Config!$E$4:$E$7", allow_blank=True)
ws_risk.add_data_validation(dv_sev)
dv_sev.add(f'E4:E{risk_end}')

dv_prob = DataValidation(type="list", formula1="Config!$G$4:$G$8", allow_blank=True)
ws_risk.add_data_validation(dv_prob)
dv_prob.add(f'F4:F{risk_end}')

dv_impact = DataValidation(type="list", formula1="Config!$E$4:$E$7", allow_blank=True)
ws_risk.add_data_validation(dv_impact)
dv_impact.add(f'G4:G{risk_end}')

dv_rstatus = DataValidation(type="list", formula1="Config!$F$4:$F$8", allow_blank=True)
ws_risk.add_data_validation(dv_rstatus)
dv_rstatus.add(f'O4:O{risk_end}')

# Summary section
sum_row = risk_end + 2
ws_risk.cell(row=sum_row, column=1, value="TỔNG HỢP RỦI RO").font = subtitle_font
ws_risk.merge_cells(f'A{sum_row}:D{sum_row}')

labels = ["Tổng số rủi ro:", "Critical:", "High:", "Open:", "Escalated:", "Risk Score TB:"]
formulas = [
    f'=COUNTA(E4:E{risk_end})',
    f'=COUNTIF(E4:E{risk_end},"Critical")',
    f'=COUNTIF(E4:E{risk_end},"High")',
    f'=COUNTIF(O4:O{risk_end},"Open")',
    f'=COUNTIF(O4:O{risk_end},"Escalated")',
    f'=AVERAGE(K4:K{risk_end})',
]
for i, (lbl, frm) in enumerate(zip(labels, formulas)):
    ws_risk.cell(row=sum_row+1+i, column=1, value=lbl).font = bold_font
    ws_risk.cell(row=sum_row+1+i, column=2).value = frm
    ws_risk.cell(row=sum_row+1+i, column=2).font = bold_font

# Risk Matrix (Heat Map style)
matrix_row = sum_row + 9
ws_risk.cell(row=matrix_row, column=1, value="RISK MATRIX (Severity x Probability)").font = subtitle_font
ws_risk.merge_cells(f'A{matrix_row}:F{matrix_row}')

# Matrix headers
ws_risk.cell(row=matrix_row+1, column=2, value="Low (1)").font = sub_header_font
ws_risk.cell(row=matrix_row+1, column=2).fill = sub_header_fill
ws_risk.cell(row=matrix_row+1, column=3, value="Medium (2)").font = sub_header_font
ws_risk.cell(row=matrix_row+1, column=3).fill = sub_header_fill
ws_risk.cell(row=matrix_row+1, column=4, value="High (3)").font = sub_header_font
ws_risk.cell(row=matrix_row+1, column=4).fill = sub_header_fill
ws_risk.cell(row=matrix_row+1, column=5, value="Critical (4)").font = sub_header_font
ws_risk.cell(row=matrix_row+1, column=5).fill = sub_header_fill

prob_labels = [("Rất cao (5)", 5), ("Cao (4)", 4), ("TB (3)", 3), ("Thấp (2)", 2), ("Rất thấp (1)", 1)]
colors_matrix = {
    (1,1): LIGHT_GREEN, (1,2): LIGHT_GREEN, (1,3): LIGHT_YELLOW, (1,4): LIGHT_YELLOW,
    (2,1): LIGHT_GREEN, (2,2): LIGHT_YELLOW, (2,3): LIGHT_YELLOW, (2,4): ORANGE,
    (3,1): LIGHT_YELLOW, (3,2): LIGHT_YELLOW, (3,3): ORANGE, (3,4): ORANGE,
    (4,1): LIGHT_YELLOW, (4,2): ORANGE, (4,3): ORANGE, (4,4): LIGHT_RED,
    (5,1): ORANGE, (5,2): ORANGE, (5,3): LIGHT_RED, (5,4): LIGHT_RED,
}

for i, (plbl, pval) in enumerate(prob_labels):
    row = matrix_row + 2 + i
    ws_risk.cell(row=row, column=1, value=plbl).font = bold_font
    ws_risk.cell(row=row, column=1).fill = header_fill
    ws_risk.cell(row=row, column=1).font = header_font
    for j, sval in enumerate([1, 2, 3, 4]):
        cell = ws_risk.cell(row=row, column=2+j)
        cell.value = pval * sval
        cell.alignment = center_align
        cell.border = thin_border
        cell.font = Font(name="Arial", bold=True, size=14)
        c = colors_matrix.get((pval, sval), WHITE)
        cell.fill = PatternFill("solid", fgColor=c)

col_widths_r = [5, 14, 25, 35, 15, 16, 15, 12, 12, 12, 12, 18, 35, 13, 13]
for i, w in enumerate(col_widths_r, 1):
    ws_risk.column_dimensions[get_column_letter(i)].width = w
ws_risk.freeze_panes = 'A4'

# ============================================================
# SHEET: Dashboard (FIRST SHEET - move to front)
# ============================================================
ws_dash = wb.create_sheet("Dashboard", 0)
ws_dash.sheet_properties.tabColor = DARK_BLUE

ws_dash['A1'] = "BÁO CÁO TIẾN ĐỘ DỰ ÁN - TIÊN PHONG CDS"
ws_dash['A1'].font = Font(name="Arial", bold=True, color=DARK_BLUE, size=20)
ws_dash.merge_cells('A1:L1')

ws_dash['A2'] = "CTO Project Status Report"
ws_dash['A2'].font = Font(name="Arial", color="666666", size=12, italic=True)
ws_dash.merge_cells('A2:L2')

# Period selector
ws_dash['A3'] = "Kỳ báo cáo:"
ws_dash['A3'].font = bold_font
ws_dash['B3'] = "Tháng 3"
ws_dash['B3'].font = Font(name="Arial", bold=True, color=MED_BLUE, size=12)
dv_month = DataValidation(type="list", formula1="Config!$A$4:$A$15", allow_blank=False)
ws_dash.add_data_validation(dv_month)
dv_month.add('B3')

ws_dash['B4'] = 2026
ws_dash['B4'].font = Font(name="Arial", bold=True, color=MED_BLUE, size=12)
dv_year = DataValidation(type="list", formula1="Config!$B$4:$B$10", allow_blank=False)
ws_dash.add_data_validation(dv_year)
dv_year.add('B4')

ws_dash['C3'] = "Ngày tạo:"
ws_dash['C3'].font = bold_font
ws_dash['D3'] = '=TODAY()'
ws_dash['D3'].number_format = date_format
ws_dash['D3'].font = bold_font

# ---- KPI CARDS ----
kpi_row = 6
ws_dash.cell(row=kpi_row, column=1, value="TỔNG QUAN DỰ ÁN").font = subtitle_font
ws_dash.merge_cells(f'A{kpi_row}:L{kpi_row}')
ws_dash.cell(row=kpi_row, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)

# KPI 1: Tổng dự án
kpi_r = kpi_row + 1
ws_dash.merge_cells(f'A{kpi_r}:B{kpi_r+1}')
ws_dash.cell(row=kpi_r, column=1, value=f'=COUNTA(\'Danh Sách Dự Án\'!B5:B11)')
ws_dash.cell(row=kpi_r, column=1).font = kpi_value_font
ws_dash.cell(row=kpi_r, column=1).alignment = center_align
ws_dash.cell(row=kpi_r, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
ws_dash.cell(row=kpi_r+2, column=1, value="Tổng Dự Án")
ws_dash.cell(row=kpi_r+2, column=1).font = kpi_label_font
ws_dash.cell(row=kpi_r+2, column=1).alignment = center_align
ws_dash.merge_cells(f'A{kpi_r+2}:B{kpi_r+2}')

# KPI 2: On Track
ws_dash.merge_cells(f'C{kpi_r}:D{kpi_r+1}')
ws_dash.cell(row=kpi_r, column=3, value=f'=COUNTIF(\'Danh Sách Dự Án\'!I5:I11,"On Track")')
ws_dash.cell(row=kpi_r, column=3).font = Font(name="Arial", bold=True, color=GREEN, size=24)
ws_dash.cell(row=kpi_r, column=3).alignment = center_align
ws_dash.cell(row=kpi_r, column=3).fill = PatternFill("solid", fgColor=LIGHT_GREEN)
ws_dash.cell(row=kpi_r+2, column=3, value="On Track")
ws_dash.cell(row=kpi_r+2, column=3).font = kpi_label_font
ws_dash.cell(row=kpi_r+2, column=3).alignment = center_align
ws_dash.merge_cells(f'C{kpi_r+2}:D{kpi_r+2}')

# KPI 3: At Risk
ws_dash.merge_cells(f'E{kpi_r}:F{kpi_r+1}')
ws_dash.cell(row=kpi_r, column=5, value=f'=COUNTIF(\'Danh Sách Dự Án\'!I5:I11,"At Risk")')
ws_dash.cell(row=kpi_r, column=5).font = Font(name="Arial", bold=True, color=ORANGE, size=24)
ws_dash.cell(row=kpi_r, column=5).alignment = center_align
ws_dash.cell(row=kpi_r, column=5).fill = PatternFill("solid", fgColor=LIGHT_YELLOW)
ws_dash.cell(row=kpi_r+2, column=5, value="At Risk")
ws_dash.cell(row=kpi_r+2, column=5).font = kpi_label_font
ws_dash.cell(row=kpi_r+2, column=5).alignment = center_align
ws_dash.merge_cells(f'E{kpi_r+2}:F{kpi_r+2}')

# KPI 4: Delayed
ws_dash.merge_cells(f'G{kpi_r}:H{kpi_r+1}')
ws_dash.cell(row=kpi_r, column=7, value=f'=COUNTIF(\'Danh Sách Dự Án\'!I5:I11,"Delayed")')
ws_dash.cell(row=kpi_r, column=7).font = Font(name="Arial", bold=True, color=RED, size=24)
ws_dash.cell(row=kpi_r, column=7).alignment = center_align
ws_dash.cell(row=kpi_r, column=7).fill = PatternFill("solid", fgColor=LIGHT_RED)
ws_dash.cell(row=kpi_r+2, column=7, value="Delayed")
ws_dash.cell(row=kpi_r+2, column=7).font = kpi_label_font
ws_dash.cell(row=kpi_r+2, column=7).alignment = center_align
ws_dash.merge_cells(f'G{kpi_r+2}:H{kpi_r+2}')

# KPI 5: Completed
ws_dash.merge_cells(f'I{kpi_r}:J{kpi_r+1}')
ws_dash.cell(row=kpi_r, column=9, value=f'=COUNTIF(\'Danh Sách Dự Án\'!I5:I11,"Completed")')
ws_dash.cell(row=kpi_r, column=9).font = Font(name="Arial", bold=True, color=DARK_GREEN, size=24)
ws_dash.cell(row=kpi_r, column=9).alignment = center_align
ws_dash.cell(row=kpi_r, column=9).fill = PatternFill("solid", fgColor="C6EFCE")
ws_dash.cell(row=kpi_r+2, column=9, value="Completed")
ws_dash.cell(row=kpi_r+2, column=9).font = kpi_label_font
ws_dash.cell(row=kpi_r+2, column=9).alignment = center_align
ws_dash.merge_cells(f'I{kpi_r+2}:J{kpi_r+2}')

# KPI 6: % Hoàn thành TB
ws_dash.merge_cells(f'K{kpi_r}:L{kpi_r+1}')
ws_dash.cell(row=kpi_r, column=11, value=f'=\'Danh Sách Dự Án\'!K12')
ws_dash.cell(row=kpi_r, column=11).font = kpi_value_font
ws_dash.cell(row=kpi_r, column=11).number_format = '0%'
ws_dash.cell(row=kpi_r, column=11).alignment = center_align
ws_dash.cell(row=kpi_r, column=11).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
ws_dash.cell(row=kpi_r+2, column=11, value="% HT Trung Bình")
ws_dash.cell(row=kpi_r+2, column=11).font = kpi_label_font
ws_dash.cell(row=kpi_r+2, column=11).alignment = center_align
ws_dash.merge_cells(f'K{kpi_r+2}:L{kpi_r+2}')

# ---- FINANCIAL KPIs ----
fin_row = kpi_r + 4
ws_dash.cell(row=fin_row, column=1, value="TỔNG QUAN TÀI CHÍNH").font = subtitle_font
ws_dash.merge_cells(f'A{fin_row}:L{fin_row}')
ws_dash.cell(row=fin_row, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)

fr = fin_row + 1
# Total contract value
ws_dash.merge_cells(f'A{fr}:C{fr+1}')
ws_dash.cell(row=fr, column=1, value=f'=\'Danh Sách Dự Án\'!M12')
ws_dash.cell(row=fr, column=1).font = Font(name="Arial", bold=True, color=DARK_BLUE, size=18)
ws_dash.cell(row=fr, column=1).number_format = '#,##0,,"Tỷ"'
ws_dash.cell(row=fr, column=1).alignment = center_align
ws_dash.cell(row=fr, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
ws_dash.cell(row=fr+2, column=1, value="Tổng Giá Trị HĐ (VNĐ)")
ws_dash.cell(row=fr+2, column=1).font = kpi_label_font
ws_dash.cell(row=fr+2, column=1).alignment = center_align
ws_dash.merge_cells(f'A{fr+2}:C{fr+2}')

# Total received
ws_dash.merge_cells(f'D{fr}:F{fr+1}')
ws_dash.cell(row=fr, column=4, value=f'=\'Danh Sách Dự Án\'!N12')
ws_dash.cell(row=fr, column=4).font = Font(name="Arial", bold=True, color=GREEN, size=18)
ws_dash.cell(row=fr, column=4).number_format = '#,##0,,"Tỷ"'
ws_dash.cell(row=fr, column=4).alignment = center_align
ws_dash.cell(row=fr, column=4).fill = PatternFill("solid", fgColor=LIGHT_GREEN)
ws_dash.cell(row=fr+2, column=4, value="Đã Thu (VNĐ)")
ws_dash.cell(row=fr+2, column=4).font = kpi_label_font
ws_dash.cell(row=fr+2, column=4).alignment = center_align
ws_dash.merge_cells(f'D{fr+2}:F{fr+2}')

# Remaining
ws_dash.merge_cells(f'G{fr}:I{fr+1}')
ws_dash.cell(row=fr, column=7, value=f'=\'Danh Sách Dự Án\'!O12')
ws_dash.cell(row=fr, column=7).font = Font(name="Arial", bold=True, color=ORANGE, size=18)
ws_dash.cell(row=fr, column=7).number_format = '#,##0,,"Tỷ"'
ws_dash.cell(row=fr, column=7).alignment = center_align
ws_dash.cell(row=fr, column=7).fill = PatternFill("solid", fgColor=LIGHT_YELLOW)
ws_dash.cell(row=fr+2, column=7, value="Còn Phải Thu (VNĐ)")
ws_dash.cell(row=fr+2, column=7).font = kpi_label_font
ws_dash.cell(row=fr+2, column=7).alignment = center_align
ws_dash.merge_cells(f'G{fr+2}:I{fr+2}')

# % Payment
ws_dash.merge_cells(f'J{fr}:L{fr+1}')
ws_dash.cell(row=fr, column=10, value=f'=\'Danh Sách Dự Án\'!P12')
ws_dash.cell(row=fr, column=10).font = kpi_value_font
ws_dash.cell(row=fr, column=10).number_format = '0%'
ws_dash.cell(row=fr, column=10).alignment = center_align
ws_dash.cell(row=fr, column=10).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
ws_dash.cell(row=fr+2, column=10, value="% Đã Thanh Toán")
ws_dash.cell(row=fr+2, column=10).font = kpi_label_font
ws_dash.cell(row=fr+2, column=10).alignment = center_align
ws_dash.merge_cells(f'J{fr+2}:L{fr+2}')

# ---- RISK & HR KPIs ----
rh_row = fr + 4
ws_dash.cell(row=rh_row, column=1, value="RỦI RO & NHÂN SỰ").font = subtitle_font
ws_dash.merge_cells(f'A{rh_row}:L{rh_row}')
ws_dash.cell(row=rh_row, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)

rr = rh_row + 1
# Total risks
ws_dash.merge_cells(f'A{rr}:C{rr+1}')
ws_dash.cell(row=rr, column=1, value=f'=\'Rủi Ro & Vấn Đề\'!B{sum_row+1}')
ws_dash.cell(row=rr, column=1).font = kpi_value_font
ws_dash.cell(row=rr, column=1).alignment = center_align
ws_dash.cell(row=rr, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
ws_dash.cell(row=rr+2, column=1, value="Tổng Rủi Ro")
ws_dash.cell(row=rr+2, column=1).font = kpi_label_font
ws_dash.cell(row=rr+2, column=1).alignment = center_align
ws_dash.merge_cells(f'A{rr+2}:C{rr+2}')

# Critical risks
ws_dash.merge_cells(f'D{rr}:F{rr+1}')
ws_dash.cell(row=rr, column=4, value=f'=\'Rủi Ro & Vấn Đề\'!B{sum_row+2}')
ws_dash.cell(row=rr, column=4).font = Font(name="Arial", bold=True, color=RED, size=24)
ws_dash.cell(row=rr, column=4).alignment = center_align
ws_dash.cell(row=rr, column=4).fill = PatternFill("solid", fgColor=LIGHT_RED)
ws_dash.cell(row=rr+2, column=4, value="Critical")
ws_dash.cell(row=rr+2, column=4).font = kpi_label_font
ws_dash.cell(row=rr+2, column=4).alignment = center_align
ws_dash.merge_cells(f'D{rr+2}:F{rr+2}')

# Total staff
ws_dash.merge_cells(f'G{rr}:I{rr+1}')
ws_dash.cell(row=rr, column=7, value=f'=\'Danh Sách Dự Án\'!Q12')
ws_dash.cell(row=rr, column=7).font = kpi_value_font
ws_dash.cell(row=rr, column=7).alignment = center_align
ws_dash.cell(row=rr, column=7).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
ws_dash.cell(row=rr+2, column=7, value="Tổng Nhân Sự")
ws_dash.cell(row=rr+2, column=7).font = kpi_label_font
ws_dash.cell(row=rr+2, column=7).alignment = center_align
ws_dash.merge_cells(f'G{rr+2}:I{rr+2}')

# Avg utilization
ws_dash.merge_cells(f'J{rr}:L{rr+1}')
ws_dash.cell(row=rr, column=10, value=f'=\'Nhân Sự Dự Án\'!H{hr_total}')
ws_dash.cell(row=rr, column=10).font = kpi_value_font
ws_dash.cell(row=rr, column=10).number_format = '0%'
ws_dash.cell(row=rr, column=10).alignment = center_align
ws_dash.cell(row=rr, column=10).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
ws_dash.cell(row=rr+2, column=10, value="Utilization Rate TB")
ws_dash.cell(row=rr+2, column=10).font = kpi_label_font
ws_dash.cell(row=rr+2, column=10).alignment = center_align
ws_dash.merge_cells(f'J{rr+2}:L{rr+2}')

# ---- CHARTS ----
chart_start = rr + 4

# Chart 1: Pie chart - Project Status
pie = PieChart()
pie.title = "Phân Bổ Trạng Thái Dự Án"
pie.style = 10
pie.width = 14
pie.height = 12

# Create temp data for pie chart on dashboard
ws_dash.cell(row=chart_start, column=1, value="Trạng Thái").font = Font(name="Arial", size=8, color="FFFFFF")
ws_dash.cell(row=chart_start+1, column=1, value="On Track").font = Font(name="Arial", size=8, color="FFFFFF")
ws_dash.cell(row=chart_start+2, column=1, value="At Risk").font = Font(name="Arial", size=8, color="FFFFFF")
ws_dash.cell(row=chart_start+3, column=1, value="Delayed").font = Font(name="Arial", size=8, color="FFFFFF")
ws_dash.cell(row=chart_start+4, column=1, value="Completed").font = Font(name="Arial", size=8, color="FFFFFF")

ws_dash.cell(row=chart_start, column=2, value="Số lượng").font = Font(name="Arial", size=8, color="FFFFFF")
ws_dash.cell(row=chart_start+1, column=2).value = f'=COUNTIF(\'Danh Sách Dự Án\'!I5:I11,"On Track")'
ws_dash.cell(row=chart_start+2, column=2).value = f'=COUNTIF(\'Danh Sách Dự Án\'!I5:I11,"At Risk")'
ws_dash.cell(row=chart_start+3, column=2).value = f'=COUNTIF(\'Danh Sách Dự Án\'!I5:I11,"Delayed")'
ws_dash.cell(row=chart_start+4, column=2).value = f'=COUNTIF(\'Danh Sách Dự Án\'!I5:I11,"Completed")'

# Hide the data by making font white
for r in range(chart_start, chart_start+5):
    for c in [1, 2]:
        ws_dash.cell(row=r, column=c).font = Font(name="Arial", size=1, color="FFFFFF")

cats_pie = Reference(ws_dash, min_col=1, min_row=chart_start+1, max_row=chart_start+4)
data_pie = Reference(ws_dash, min_col=2, min_row=chart_start, max_row=chart_start+4)
pie.add_data(data_pie, titles_from_data=True)
pie.set_categories(cats_pie)

# Colors for pie
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
pie_colors = [GREEN, ORANGE, RED, DARK_GREEN]
for i, color in enumerate(pie_colors):
    pt = DataPoint(idx=i)
    pt.graphicalProperties.solidFill = color
    pie.series[0].data_points.append(pt)

pie.dataLabels = DataLabelList()
pie.dataLabels.showPercent = True
pie.dataLabels.showVal = True
ws_dash.add_chart(pie, f"A{chart_start+5}")

# Chart 2: Bar chart - Progress comparison
bar = BarChart()
bar.type = "col"
bar.grouping = "clustered"
bar.title = "Tiến Độ: Kế Hoạch vs Thực Tế"
bar.y_axis.title = "% Hoàn Thành"
bar.y_axis.scaling.max = 1.0
bar.style = 10
bar.width = 20
bar.height = 12

# Data from project sheet
cats_bar = Reference(ws_projects, min_col=3, min_row=5, max_row=11)
data_planned = Reference(ws_projects, min_col=10, min_row=4, max_row=11)
data_actual = Reference(ws_projects, min_col=11, min_row=4, max_row=11)
bar.add_data(data_planned, titles_from_data=True)
bar.add_data(data_actual, titles_from_data=True)
bar.set_categories(cats_bar)
bar.series[0].graphicalProperties.solidFill = LIGHT_BLUE
bar.series[1].graphicalProperties.solidFill = MED_BLUE
ws_dash.add_chart(bar, f"G{chart_start+5}")

# Chart 3: Revenue chart
bar_rev = BarChart()
bar_rev.type = "col"
bar_rev.grouping = "stacked"
bar_rev.title = "Doanh Thu Hợp Đồng: Đã Thu vs Còn Phải Thu"
bar_rev.y_axis.title = "VNĐ"
bar_rev.style = 10
bar_rev.width = 20
bar_rev.height = 12

cats_rev = Reference(ws_projects, min_col=3, min_row=5, max_row=11)
data_paid_rev = Reference(ws_projects, min_col=14, min_row=4, max_row=11)
data_remain_rev = Reference(ws_projects, min_col=15, min_row=4, max_row=11)
bar_rev.add_data(data_paid_rev, titles_from_data=True)
bar_rev.add_data(data_remain_rev, titles_from_data=True)
bar_rev.set_categories(cats_rev)
bar_rev.series[0].graphicalProperties.solidFill = GREEN
bar_rev.series[1].graphicalProperties.solidFill = ORANGE
ws_dash.add_chart(bar_rev, f"A{chart_start+22}")

# Chart 4: Utilization chart
bar_util_dash = BarChart()
bar_util_dash.type = "col"
bar_util_dash.title = "Utilization Rate Nhân Sự"
bar_util_dash.y_axis.title = "Rate"
bar_util_dash.y_axis.scaling.max = 1.5
bar_util_dash.style = 10
bar_util_dash.width = 20
bar_util_dash.height = 12

cats_util_d = Reference(ws_hr, min_col=2, min_row=4, max_row=hr_total-1)
data_util_d = Reference(ws_hr, min_col=8, min_row=3, max_row=hr_total-1)
bar_util_dash.add_data(data_util_d, titles_from_data=True)
bar_util_dash.set_categories(cats_util_d)
bar_util_dash.series[0].graphicalProperties.solidFill = MED_BLUE
ws_dash.add_chart(bar_util_dash, f"G{chart_start+22}")

# Column widths for dashboard
for i in range(1, 13):
    ws_dash.column_dimensions[get_column_letter(i)].width = 14

ws_dash.sheet_view.zoomScale = 90

# ============================================================
# SHEET: Báo Cáo Tháng
# ============================================================
ws_monthly = wb.create_sheet("Báo Cáo Tháng")
ws_monthly.sheet_properties.tabColor = "7030A0"

ws_monthly['A1'] = "BÁO CÁO TỔNG HỢP THÁNG"
ws_monthly['A1'].font = title_font
ws_monthly.merge_cells('A1:H1')

ws_monthly['A2'] = "Kỳ:"
ws_monthly['A2'].font = bold_font
ws_monthly['B2'] = "Tháng 3"
dv_m2 = DataValidation(type="list", formula1="Config!$A$4:$A$15", allow_blank=False)
ws_monthly.add_data_validation(dv_m2)
dv_m2.add('B2')
ws_monthly['C2'] = "Năm:"
ws_monthly['C2'].font = bold_font
ws_monthly['D2'] = 2026
dv_y2 = DataValidation(type="list", formula1="Config!$B$4:$B$10", allow_blank=False)
ws_monthly.add_data_validation(dv_y2)
dv_y2.add('D2')

# Section 1: Summary
s1 = 4
ws_monthly.cell(row=s1, column=1, value="1. TỔNG QUAN DỰ ÁN").font = subtitle_font
ws_monthly.merge_cells(f'A{s1}:H{s1}')
ws_monthly.cell(row=s1, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)

metrics = [
    ("Tổng số dự án đang triển khai", f'=COUNTIFS(\'Danh Sách Dự Án\'!I5:I11,"<>Completed")'),
    ("Dự án On Track", f'=COUNTIF(\'Danh Sách Dự Án\'!I5:I11,"On Track")'),
    ("Dự án At Risk", f'=COUNTIF(\'Danh Sách Dự Án\'!I5:I11,"At Risk")'),
    ("Dự án Delayed", f'=COUNTIF(\'Danh Sách Dự Án\'!I5:I11,"Delayed")'),
    ("Dự án Completed", f'=COUNTIF(\'Danh Sách Dự Án\'!I5:I11,"Completed")'),
    ("% Dự án On Track", f'=IFERROR(COUNTIF(\'Danh Sách Dự Án\'!I5:I11,"On Track")/COUNTA(\'Danh Sách Dự Án\'!I5:I11),0)'),
    ("% Hoàn thành trung bình", f'=\'Danh Sách Dự Án\'!K12'),
]
for i, (label, formula) in enumerate(metrics):
    ws_monthly.cell(row=s1+1+i, column=1, value=label).font = normal_font
    ws_monthly.cell(row=s1+1+i, column=2).value = formula
    ws_monthly.cell(row=s1+1+i, column=2).font = bold_font
    if '%' in label:
        ws_monthly.cell(row=s1+1+i, column=2).number_format = pct_format
    style_data_row(ws_monthly, s1+1+i, 2, alt=(i % 2 == 1))

# Section 2: Financial
s2 = s1 + len(metrics) + 2
ws_monthly.cell(row=s2, column=1, value="2. TÀI CHÍNH").font = subtitle_font
ws_monthly.merge_cells(f'A{s2}:H{s2}')
ws_monthly.cell(row=s2, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)

fin_metrics = [
    ("Tổng giá trị hợp đồng", f'=\'Danh Sách Dự Án\'!M12', vnd_format),
    ("Tổng đã thu", f'=\'Danh Sách Dự Án\'!N12', vnd_format),
    ("Tổng còn phải thu", f'=\'Danh Sách Dự Án\'!O12', vnd_format),
    ("% Đã thu", f'=\'Danh Sách Dự Án\'!P12', pct_format),
]
for i, (label, formula, fmt) in enumerate(fin_metrics):
    ws_monthly.cell(row=s2+1+i, column=1, value=label).font = normal_font
    ws_monthly.cell(row=s2+1+i, column=2).value = formula
    ws_monthly.cell(row=s2+1+i, column=2).font = bold_font
    ws_monthly.cell(row=s2+1+i, column=2).number_format = fmt
    style_data_row(ws_monthly, s2+1+i, 2, alt=(i % 2 == 1))

# Section 3: Risk Summary
s3 = s2 + len(fin_metrics) + 2
ws_monthly.cell(row=s3, column=1, value="3. RỦI RO CẦN XỬ LÝ").font = subtitle_font
ws_monthly.merge_cells(f'A{s3}:H{s3}')
ws_monthly.cell(row=s3, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)

risk_metrics = [
    ("Tổng số rủi ro đang mở", f'=COUNTIFS(\'Rủi Ro & Vấn Đề\'!O4:O11,"<>Resolved",\'Rủi Ro & Vấn Đề\'!O4:O11,"<>Closed")'),
    ("Rủi ro Critical", f'=COUNTIF(\'Rủi Ro & Vấn Đề\'!E4:E11,"Critical")'),
    ("Rủi ro đã Escalated", f'=COUNTIF(\'Rủi Ro & Vấn Đề\'!O4:O11,"Escalated")'),
    ("Risk Score trung bình", f'=AVERAGE(\'Rủi Ro & Vấn Đề\'!K4:K11)'),
]
for i, (label, formula) in enumerate(risk_metrics):
    ws_monthly.cell(row=s3+1+i, column=1, value=label).font = normal_font
    ws_monthly.cell(row=s3+1+i, column=2).value = formula
    ws_monthly.cell(row=s3+1+i, column=2).font = bold_font
    style_data_row(ws_monthly, s3+1+i, 2, alt=(i % 2 == 1))

# Section 4: HR
s4 = s3 + len(risk_metrics) + 2
ws_monthly.cell(row=s4, column=1, value="4. NHÂN SỰ").font = subtitle_font
ws_monthly.merge_cells(f'A{s4}:H{s4}')
ws_monthly.cell(row=s4, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)

hr_metrics = [
    ("Tổng nhân sự tham gia DA", f'=\'Danh Sách Dự Án\'!Q12'),
    ("Utilization Rate trung bình", f'=\'Nhân Sự Dự Án\'!H{hr_total}', pct_format),
    ("Nhân sự overloaded (>110%)", f'=COUNTIF(\'Nhân Sự Dự Án\'!H4:H{hr_total-1},">"&1.1)'),
]
for i, item in enumerate(hr_metrics):
    label = item[0]
    formula = item[1]
    fmt = item[2] if len(item) > 2 else None
    ws_monthly.cell(row=s4+1+i, column=1, value=label).font = normal_font
    ws_monthly.cell(row=s4+1+i, column=2).value = formula
    ws_monthly.cell(row=s4+1+i, column=2).font = bold_font
    if fmt:
        ws_monthly.cell(row=s4+1+i, column=2).number_format = fmt
    style_data_row(ws_monthly, s4+1+i, 2, alt=(i % 2 == 1))

# Section 5: Highlights & Action Items (manual input)
s5 = s4 + len(hr_metrics) + 2
ws_monthly.cell(row=s5, column=1, value="5. HIGHLIGHTS").font = subtitle_font
ws_monthly.merge_cells(f'A{s5}:H{s5}')
ws_monthly.cell(row=s5, column=1).fill = PatternFill("solid", fgColor=LIGHT_GREEN)
for i in range(1, 6):
    ws_monthly.cell(row=s5+i, column=1, value=f"{i}.")
    ws_monthly.merge_cells(f'B{s5+i}:H{s5+i}')
    ws_monthly.cell(row=s5+i, column=2, value="[Nhập highlight]" if i <= 3 else "")
    ws_monthly.cell(row=s5+i, column=2).font = Font(name="Arial", color="999999", size=10, italic=True)

s6 = s5 + 7
ws_monthly.cell(row=s6, column=1, value="6. LOWLIGHTS / VẤN ĐỀ").font = subtitle_font
ws_monthly.merge_cells(f'A{s6}:H{s6}')
ws_monthly.cell(row=s6, column=1).fill = PatternFill("solid", fgColor=LIGHT_RED)
for i in range(1, 6):
    ws_monthly.cell(row=s6+i, column=1, value=f"{i}.")
    ws_monthly.merge_cells(f'B{s6+i}:H{s6+i}')
    ws_monthly.cell(row=s6+i, column=2, value="[Nhập vấn đề]" if i <= 3 else "")
    ws_monthly.cell(row=s6+i, column=2).font = Font(name="Arial", color="999999", size=10, italic=True)

s7 = s6 + 7
ws_monthly.cell(row=s7, column=1, value="7. ACTION ITEMS").font = subtitle_font
ws_monthly.merge_cells(f'A{s7}:H{s7}')
ws_monthly.cell(row=s7, column=1).fill = PatternFill("solid", fgColor=LIGHT_YELLOW)

action_headers = ["STT", "Action Item", "Người Phụ Trách", "Deadline", "Trạng Thái"]
for i, h in enumerate(action_headers, 1):
    ws_monthly.cell(row=s7+1, column=i, value=h)
style_header_row(ws_monthly, s7+1, len(action_headers))
for i in range(1, 6):
    ws_monthly.cell(row=s7+1+i, column=1, value=i)
    for c in range(1, len(action_headers)+1):
        ws_monthly.cell(row=s7+1+i, column=c).border = thin_border

ws_monthly.column_dimensions['A'].width = 35
ws_monthly.column_dimensions['B'].width = 45
for c in range(3, 9):
    ws_monthly.column_dimensions[get_column_letter(c)].width = 18

# ============================================================
# Reorder sheets
# ============================================================
# Order: Dashboard, Danh Sách Dự Án, Tiến Độ Hợp Đồng, Nhân Sự Dự Án, Rủi Ro & Vấn Đề, Báo Cáo Tháng, Config
wb.move_sheet("Config", offset=6)

# ============================================================
# SAVE
# ============================================================
output_path = "/Users/mac/Documents/Work/Projects/TIEN PHONG CDS/cto-ai-assistant/.claude/drafts/CTO-Project-Status-Report-Template.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
